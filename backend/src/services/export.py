import datetime
import io
from typing import Optional

from openpyxl import Workbook as Libro
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import rpe
from models import (
    Block, Exercise, ExerciseDefinition, SetLog, User, Weekday, Workout,
)

DIAS = ["Lunes", "Martes", "Miercoles", "Jueves",
        "Viernes", "Sabado", "Domingo"]

_TINTA = "1F2933"
_CABECERA = PatternFill("solid", fgColor="C1272D")
_BANDA = PatternFill("solid", fgColor="F4F5F7")
_BLANCA = Font(bold=True, color="FFFFFF", size=11)
_NEGRITA = Font(bold=True, color=_TINTA)
_LINEA = Side(style="thin", color="D9DEE3")
_BORDE = Border(bottom=_LINEA)


def _uno_rm(log: SetLog) -> Optional[float]:
    # La tabla de RPE si el atleta la anoto; si no, Epley.
    if log.weight is None or log.reps < 1:
        return None
    if log.rpe is not None:
        try:
            return rpe.estimated_1rm(log.weight, log.reps, log.rpe)
        except ValueError:
            pass
    return log.estimated_1rm


def _tonelaje(log: SetLog) -> Optional[float]:
    if log.weight is None:
        return None
    return round(log.weight * log.reps, 1)


def _encabezar(hoja, titulos: list[str], anchos: list[int]) -> None:
    hoja.append(titulos)
    for i, ancho in enumerate(anchos, 1):
        hoja.column_dimensions[get_column_letter(i)].width = ancho
        celda = hoja.cell(row=1, column=i)
        celda.fill = _CABECERA
        celda.font = _BLANCA
        celda.alignment = Alignment(horizontal="center", vertical="center")
    hoja.freeze_panes = "A2"
    hoja.row_dimensions[1].height = 22


def _rayar(hoja) -> None:
    for fila in range(2, hoja.max_row + 1):
        for celda in hoja[fila]:
            celda.border = _BORDE
            if fila % 2 == 0:
                celda.fill = _BANDA


def _hoja_resumen(libro, bloque, atleta, entrenos, series) -> None:
    hoja = libro.active
    hoja.title = "Resumen"
    hoja.column_dimensions["A"].width = 24
    hoja.column_dimensions["B"].width = 38

    hechas = [s for s in series if s.completed_at is not None]
    volumen = sum(t for s in hechas if (t := _tonelaje(s)) is not None)

    filas = [
        ("Atleta", atleta.name),
        ("Bloque", bloque.name),
        ("Semanas", bloque.total_weeks),
        ("Inicio", bloque.start_date),
        ("Fin", bloque.end_date),
        ("Estado", bloque.status.value),
        ("", ""),
        ("Sesiones", len(entrenos)),
        ("Series planificadas", len(series)),
        ("Series realizadas", len(hechas)),
        ("Tonelaje total (kg)", round(volumen, 1)),
        ("", ""),
        ("Exportado", datetime.date.today()),
    ]
    for etiqueta, valor in filas:
        hoja.append([etiqueta, valor])
        if etiqueta:
            hoja.cell(row=hoja.max_row, column=1).font = _NEGRITA
    if bloque.notes:
        hoja.append(["", ""])
        hoja.append(["Notas", bloque.notes])
        hoja.cell(row=hoja.max_row, column=1).font = _NEGRITA
        hoja.cell(row=hoja.max_row, column=2).alignment = Alignment(wrap_text=True)


def _hoja_series(libro, bloque, entrenos, ejercicios, nombres, series) -> None:
    hoja = libro.create_sheet("Series")
    _encabezar(
        hoja,
        ["Semana", "Dia", "Fecha", "Sesion", "Ejercicio", "Serie",
         "Peso (kg)", "Reps", "RPE", "1RM est.", "Tonelaje", "Hecha"],
        [8, 11, 12, 22, 28, 7, 10, 7, 7, 10, 10, 8],
    )
    por_ejercicio = {e.id: e for e in ejercicios}
    por_entreno = {w.id: w for w in entrenos}

    for log in series:
        ejercicio = por_ejercicio.get(log.exercise_id)
        if ejercicio is None:
            continue
        entreno = por_entreno.get(ejercicio.workout_id)
        if entreno is None:
            continue
        hoja.append([
            entreno.week_number,
            DIAS[int(entreno.day_of_week)],
            bloque.date_for(entreno.week_number, entreno.day_of_week),
            entreno.name,
            nombres.get(ejercicio.definition_id, "?"),
            log.set_number,
            log.weight,
            log.reps,
            log.rpe,
            _uno_rm(log),
            _tonelaje(log),
            "Si" if log.completed_at else "Pendiente",
        ])
    _rayar(hoja)


def _hoja_progreso(libro, bloque, entrenos, ejercicios, nombres, series) -> None:
    hoja = libro.create_sheet("Progreso")
    semanas = list(range(1, bloque.total_weeks + 1))
    _encabezar(
        hoja,
        ["Ejercicio", "Metrica"] + [f"S{n}" for n in semanas],
        [28, 14] + [8] * len(semanas),
    )
    por_ejercicio = {e.id: e for e in ejercicios}
    por_entreno = {w.id: w for w in entrenos}

    # {definition_id: {semana: [logs]}}
    tabla: dict[int, dict[int, list[SetLog]]] = {}
    for log in series:
        if log.completed_at is None:
            continue
        ejercicio = por_ejercicio.get(log.exercise_id)
        entreno = por_entreno.get(ejercicio.workout_id) if ejercicio else None
        if entreno is None:
            continue
        semana = tabla.setdefault(ejercicio.definition_id, {})
        semana.setdefault(entreno.week_number, []).append(log)

    for definicion in sorted(tabla, key=lambda d: nombres.get(d, "")):
        por_semana = tabla[definicion]
        nombre = nombres.get(definicion, "?")

        mejores, volumenes = [], []
        for n in semanas:
            logs = por_semana.get(n, [])
            marcas = [v for log in logs if (v := _uno_rm(log)) is not None]
            cargas = [v for log in logs if (v := _tonelaje(log)) is not None]
            mejores.append(max(marcas) if marcas else None)
            volumenes.append(round(sum(cargas), 1) if cargas else None)

        hoja.append([nombre, "1RM est."] + mejores)
        hoja.cell(row=hoja.max_row, column=1).font = _NEGRITA
        hoja.append(["", "Tonelaje"] + volumenes)
    _rayar(hoja)


def libro_de_bloque(
    bloque: Block,
    atleta: User,
    entrenos: list[Workout],
    ejercicios: list[Exercise],
    definiciones: list[ExerciseDefinition],
    series: list[SetLog],
) -> bytes:
    nombres = {d.id: d.name for d in definiciones}
    libro = Libro()
    _hoja_resumen(libro, bloque, atleta, entrenos, series)
    _hoja_series(libro, bloque, entrenos, ejercicios, nombres, series)
    _hoja_progreso(libro, bloque, entrenos, ejercicios, nombres, series)

    memoria = io.BytesIO()
    libro.save(memoria)
    return memoria.getvalue()


def nombre_de_fichero(bloque: Block, atleta: User) -> str:
    limpio = "".join(
        c if c.isalnum() or c in " -_" else "" for c in f"{atleta.name} {bloque.name}"
    )
    return f"{limpio.strip().replace(' ', '_')}.xlsx"
